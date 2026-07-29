#!/usr/bin/env python3
# =============================================================================
# BUILD ISSUE CSV
# =============================================================================
# Generated: 2026-07-29 11:05 AM EDT
#
# WHAT THIS SCRIPT DOES
# ----------------------------------------------------------------------------
# You have two things:
#
#   1. INFORMATION ABOUT YOUR TIFF FILES - one TIFF per newspaper *page*,
#      named like:  18850116_vXI_n41_0001.tif
#                   |________| |_| |__| |__|
#                    date       vol issue page
#
#      This can come from any ONE of three sources (auto-detected):
#        (a) a real folder full of .tif files on your computer, OR
#        (b) a CSV file listing one full TIFF path per line, e.g.:
#              /opt/ingest_data/uploads/3/1885/18850102_vXI_n39_0001.tif
#        (c) a text file that's a Windows "dir" command listing, e.g.:
#              06/24/2025  08:05 AM   9,082,646  18850102_vXI_n39_0001.tif
#
#   2. A "TOWN LIST" FILE - a spreadsheet (xlsx, csv, or a Google Sheets
#      URL) that a human filled in by looking at each issue and writing
#      down which towns that issue covers. This file can be in EITHER of
#      two layouts (we auto-detect which one you gave us):
#
#        FORMAT A ("old method" - one row per issue):
#            filename                 | No. | Subject 1  | Subject 2 | ...
#            18850102_vXI_n39.pdf     | 39  | Deep River | Chester   | ...
#
#        FORMAT B ("new method" - one column per issue, towns stacked down):
#            Paper No. | 39         | 40         | ...
#            Towns     | Deep River | Deep River | ...
#                       | Chester    | Chester    | ...
#
# WHAT THIS SCRIPT PRODUCES (written to the same folder as the TIFF
# source - the folder itself if you gave us one, or the folder the
# listing file lives in if you gave us that instead)
# ----------------------------------------------------------------------------
#   <base_name>_metadata.csv
#       This is the FULL, 64-COLUMN, ingest-ready CSV - the same shape
#       used elsewhere in the pipeline (process_newspapers.py):
#
#         - One "Publication Issue" PARENT row per issue, with fields
#           like held_by, rights_statement, description, persons,
#           origin_information, genre, notes, etc. already filled in
#           with the project's fixed values (see CONFIGURATION below).
#           subject / geographic_subject on this row are auto-filled
#           from the Town List (see below).
#
#         - One "Page" CHILD row per TIFF page, linked back to its
#           parent issue via the member_of column, with its own
#           digital_file path and page-number-based weight/title.
#
#       "ID" numbers count up continuously across the WHOLE file (parent
#       and page rows share one counter), matching how the repository
#       ingest format expects rows to be linked.
#
#       subject is auto-filled as:      "Newspaper -- <Town>, CT"  (one
#                                        per town, joined with ^^)
#       geographic_subject is filled as: "<Town> (Conn.)"           (one
#                                        per town, joined with ^^)
#
#       If an issue has TIFFs but no matching Town List entry, its
#       subject / geographic_subject are simply left blank - the issue
#       still gets a full row (so nothing is silently dropped from the
#       ingest file), but the errors CSV will flag it so you know to
#       track down the missing town data.
#
#   <base_name>_errors.csv
#       Every mismatch we found, so you can go fix the source data:
#           missing_tiffs       -> the Town List has a paper number that
#                                  doesn't have any matching TIFF files
#                                  (this paper number will NOT appear in
#                                  the metadata CSV at all, since there's
#                                  nothing to scan)
#           missing_town_list   -> we found TIFFs for a paper number that
#                                  isn't in the Town List (this issue DOES
#                                  still appear in the metadata CSV, just
#                                  with blank subject/geographic_subject)
#
# HOW TO RUN
# ----------------------------------------------------------------------------
#   python build_issue_csv.py /path/to/tiff_directory /path/to/TownList.xlsx
#
#   ...or use a listing file instead of a real folder:
#
#   python build_issue_csv.py /path/to/tiff_list.csv /path/to/TownList.xlsx
#
#   ...and/or a Google Sheets URL instead of a local Town List file:
#
#   python build_issue_csv.py /path/to/tiff_directory "https://docs.google.com/spreadsheets/d/XXXXXXXX/edit"
#
#   For a Google Sheets URL to work, the sheet's sharing setting must be
#   at least "Anyone with the link can view" (Share button, top right of
#   the sheet, then "Change to anyone with the link"). No Google login
#   or API key is needed - the script just downloads it as a CSV.
#
# DEPENDENCIES
# ----------------------------------------------------------------------------
#   pip install openpyxl --break-system-packages
#   (openpyxl is only needed if your Town List file is .xlsx; plain .csv
#    Town Lists, Google Sheets URLs, and TIFF listing files need no extra
#    install at all)
# =============================================================================

import os
import re
import io
import sys
import csv
import urllib.request
import urllib.error
from datetime import datetime

# =============================================================================
# CONFIGURATION - fixed values that go into every "Publication Issue" row.
# These match the values already established for process_newspapers.py -
# edit here if any of them ever change.
# =============================================================================

HELD_BY                      = "Deep River Public Library"
RIGHTS_STATEMENT             = "NO COPYRIGHT - UNITED STATES"
PUBLISH                      = "Y"
RESOURCE_TYPE                = "Text"
NEWSPAPER_NAME                = "Deep River New Era"
DESCRIPTION = ("Contact the Deep River Historical Society at "
               "860-526-1449 or info@deepriverhistoricalsociety.org.")

# member_of_existing_entity_id - the parent collection ID in the repository
MEMBER_OF_EXISTING_ENTITY_ID = 61106

GENRE = "newspapers"

# persons field (column T) - the publisher's name changes based on the
# issue's year (see PUBLISHER_BY_DATE below); everything else is fixed.
PERSONS_STATIC_SUFFIX = (
    "Deep River Historical Society|relators:ctb^^"
    "Webb, Robert; Charter Oak Scanning and Digitization|relators:ctb"
)

# notes field (column AB) - fixed credits block for every issue, loaded
# from NOTES_CONFIG_FILENAME (a plain text file next to this script) so
# it can be updated by editing that file directly - no code changes
# needed. See load_notes() below for the file format and location.
NOTES_CONFIG_FILENAME = "notes_config.txt"

_notes_cache = None  # filled in by load_notes() the first time it's called


def get_notes_config_path():
    """The notes config file always lives in the SAME FOLDER as this
    script itself (not wherever the data files happen to be), so it
    stays with the code rather than getting mixed in with Google Drive
    data folders."""
    return os.path.join(
        os.path.dirname(os.path.abspath(__file__)), NOTES_CONFIG_FILENAME
    )


def load_notes():
    """
    Read the notes config file and build the notes field's value from
    it. The file format is simple - one credit per line:

        note_type: note text

    e.g.:
        creation/production credits: Deep River Public Library -- Patrick McGlamery
        creation/production credits: Chester Historical Society -- Adrian Nicholls

    Blank lines and lines starting with "#" are ignored, so the file can
    have comments/spacing for readability. The result is cached after
    the first read, so the file is only read once per run even though
    every issue needs this value.
    """
    global _notes_cache
    if _notes_cache is not None:
        return _notes_cache

    config_path = get_notes_config_path()
    if not os.path.isfile(config_path):
        raise FileNotFoundError(
            f"Notes config file not found: {config_path}\n"
            f"Create it with one 'note_type: note text' entry per line, e.g.:\n"
            f"  creation/production credits: Deep River Public Library -- Patrick McGlamery"
        )

    entries = []
    with open(config_path, "r", encoding="utf-8") as f:
        for line_number, raw_line in enumerate(f, start=1):
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if ":" not in line:
                print(f"WARNING: {config_path} line {line_number} has no "
                      f"':' separator, skipping: {line}")
                continue
            note_type, _, note = line.partition(":")
            entries.append(f"{note_type.strip()}|{note.strip()}")

    if not entries:
        raise ValueError(f"Notes config file has no valid entries: {config_path}")

    _notes_cache = MULTI_SEP.join(entries)
    return _notes_cache

# origin_information - the parts of this field that are the SAME on every
# issue. date_created (from the TIFF filename) and date_captured (today's
# date, at the moment this script runs) are filled in per-issue by
# build_origin_information() below.
ORIGIN_PLACE = "Deep River (Conn.)"
ORIGIN_ISSUANCE = "serial"
ORIGIN_FREQUENCY = "weekly"


def build_origin_information(date_created, publisher):
    """
    Build one issue's origin_information value.

    Field order (13 pipe-delimited subfields):
        event_type|place|date_created|date_issued|date_captured|
        date_valid|date_modified|other_date|copyright_date|
        publisher|edition|issuance|frequency

    date_created  - the issue's publication date (EDTF, e.g. "1885-01-16"),
                    parsed from the TIFF filename.
    publisher     - passed in rather than hardcoded, since the publisher
                    name changes depending on the issue's date (see
                    PUBLISHER_BY_DATE below).
    date_captured - always "today" (the date THIS SCRIPT is run), not a
                    fixed value, since it records when digitization
                    happened for whichever batch is being processed now.
    """
    date_captured = datetime.now().strftime("%Y-%m-%d")
    fields = [
        "",                 # event_type
        ORIGIN_PLACE,       # place
        date_created,       # date_created
        "",                 # date_issued
        date_captured,      # date_captured
        "",                 # date_valid
        "",                 # date_modified
        "",                 # other_date
        "",                 # copyright_date
        publisher,          # publisher
        "",                 # edition
        ORIGIN_ISSUANCE,    # issuance
        ORIGIN_FREQUENCY,   # frequency
    ]
    return "|".join(fields)


# The publisher listed in the persons field (column T) is the PERSON who
# published the paper at the time of that issue - looked up by the
# issue's date_created (from the TIFF filename). Names are stored
# "Last, First Middle" since that's the format the persons field needs.
#
# ASSUMPTION: where two ranges share a year (e.g. Kirtland's "...1903" and
# Prann's "1903..."), we treat the LATER publisher as starting Jan 1 of
# that shared year, since we don't have an exact transition date. Flag
# this if you ever find out the actual changeover dates.
PUBLISHER_BY_DATE = [
    ("1874-01-01", "Sheldon, Francis"),
    ("1885-01-01", "Kirtland, Charles A."),
    ("1903-01-01", "Prann, Ernest L."),
    ("1946-01-01", "Johnson, Curtiss S."),
]


def get_publisher(date_created):
    """
    Look up which publisher was in charge on a given issue's date_created
    (a "YYYY-MM-DD" string). Dates compare correctly as plain strings as
    long as they're all in that form, so no date-parsing is needed.
    """
    publisher = PUBLISHER_BY_DATE[0][1]
    for effective_date, name in PUBLISHER_BY_DATE:
        if date_created >= effective_date:
            publisher = name
        else:
            break
    return publisher


def build_persons(date_created):
    """
    Build one issue's persons value: the era-appropriate publisher
    (relators:pbl), followed by the fixed credits (relators:ctb).
    """
    publisher = get_publisher(date_created)
    return f"{publisher}|relators:pbl^^{PERSONS_STATIC_SUFFIX}"

# The server-side path BASE (no trailing subfolder) that digital_file
# values are built from. The subfolder is always the issue's YEAR (e.g.
# "1885"), per the decision to standardize on year-based folders on the
# server going forward - even if a listing file's own text shows an old
# volume-based path (e.g. ".../3/XI/..."), we build the year-based path
# ourselves rather than trust what's in the file. (Any files that are
# still sitting in old-style folders on the server will need a separate,
# later cleanup pass - not handled by this script.)
DIGITAL_FILE_BASE = "/opt/ingest_data/uploads/3/"


def build_digital_file(date_iso, filename):
    """
    e.g. build_digital_file("1885-01-16", "18850116_vXI_n41_0001.tif")
      -> "/opt/ingest_data/uploads/3/1885/18850116_vXI_n41_0001.tif"
    """
    year = date_iso[:4]
    return f"{DIGITAL_FILE_BASE}{year}/{filename}"

# The character(s) used to glue multiple values into one CSV cell.
MULTI_SEP = "^^"

# Ingest CSVs should stay well under the repository's hard limit of 1000
# lines (including the header). We roll over to a new "part" file once
# adding the NEXT issue would push a file past this target - so files
# land in the ~750-800 line range in practice, and a file is never split
# in the middle of an issue (an issue's parent + page rows always stay
# together in one file).
TARGET_LINES_PER_FILE = 800

# When issues are more than this many years apart, force a new output
# file rather than lumping a big historical gap into one file (e.g. an
# 1876 issue immediately followed by an 1885 issue, with nothing in
# between). A gap of 1 is normal (a volume straddling New Year's), so
# this only kicks in for genuinely non-contiguous data.
MAX_CONTIGUOUS_YEAR_GAP = 1

# All 64 column names, in the exact order the repository ingest expects.
COLUMNS = [
    "ID", "member_of", "member_of_existing_entity_id", "publish",
    "model", "held_by", "title", "rights_statement", "weight",
    "digital_file", "media_use", "digital_origin", "creative_commons",
    "use_and_reproduction", "restriction_on_access", "resource_type",
    "description", "local_identifier", "family_name", "persons",
    "organizations", "origin_information", "language", "genre",
    "subject", "temporal_subject", "geographic_subject", "notes",
    "scale", "physical_form", "reformatting_quality",
    "physical_description_note", "access_restriction_term",
    "related_items", "table_of_contents", "handle", "isbn",
    "oclc_number", "loc_classification", "coordinates",
    "geographic_code", "west_bounding_coordinate",
    "east_bounding_coordinate", "north_bounding_coordinate",
    "south_bounding_coordinate", "extent", "physical_location",
    "sublocation", "shelf_locator", "location_url",
    "record_information", "degree_name", "degree_level",
    "degree_discipline", "kingdom_dwr", "phylum_dwr", "class_dwr",
    "order_dwr", "family_dwr", "genus_dwr", "sci_name_auth_dwr",
    "locality_dwr", "location_remarks_dwr", "recorded_by_dwr",
]

# The pattern every TIFF filename must match. Written out piece by piece:
#   (\d{8})       - 8 digit date, e.g. 18850116
#   _v([A-Za-z0-9]+) - underscore, "v", then the volume - normally roman
#                      numerals ("XI"), but some early files use a plain
#                      number ("1") instead - both are accepted here and
#                      normalized to roman numerals in parse_tiff_filename
#   _n(\d+)          - underscore, "n", then the issue/paper number
#   (?:[_-](\d+))?   - OPTIONAL underscore-OR-hyphen + page number, any
#                      number of digits (usually "_0001", but some early
#                      files use "-000001" instead - a hyphen and/or
#                      extra leading zeros)
#   \.tiff?          - ".tif" or ".tiff", case-insensitive
TIFF_NAME_PATTERN = re.compile(
    r'^(\d{8})_v([A-Za-z0-9]+)_n(\d+)(?:[_-](\d+))?\.tiff?$',
    re.IGNORECASE,
)

# Same idea as TIFF_NAME_PATTERN above, but WITHOUT the ^ and $ anchors,
# so it can find a TIFF filename ANYWHERE inside a longer line of text -
# e.g. buried in a full path ("/opt/.../18850116_vXI_n41_0001.tif") or
# at the end of a directory-listing line ("06/24/2025 8:05 AM  9,082,646
# 18850116_vXI_n41_0001.tif"). Used by read_tiff_list_file() below.
TIFF_NAME_SEARCH_PATTERN = re.compile(
    r'(\d{8}_v[A-Za-z0-9]+_n\d+(?:[_-]\d+)?\.tiff?)',
    re.IGNORECASE,
)

# Matches the sheet ID out of any Google Sheets URL, e.g. out of:
#   https://docs.google.com/spreadsheets/d/1AbCdEfGhIjKlMnOp/edit#gid=0
# this captures "1AbCdEfGhIjKlMnOp".
GOOGLE_SHEETS_URL_PATTERN = re.compile(
    r'docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)'
)

# Matches the specific TAB (worksheet) of the sheet, if the URL includes
# one, e.g. the "0" in "...edit#gid=0". If a Town List spreadsheet has
# more than one tab, this is how we know which one to read.
GOOGLE_SHEETS_GID_PATTERN = re.compile(r'[#&?]gid=(\d+)')

# The two kinds of input files. Rather than requiring a specific prefix
# grammar, these just look for the "tifflist" / "town list" keyword
# ANYWHERE in the filename (allowing .txt or .csv for the tifflist,
# .xlsx or .csv for the Town List) - e.g. "Vol_09_tifflist.txt",
# "vI_tifflist.txt", "09_tifflist.txt", "Vol-1-Town-List.xlsx", and
# "Town_List_1885_Old_Method_-_1885.csv" (keyword NOT at the end) all
# get recognized. extract_volume_token() + normalize_volume() below turn
# whatever text surrounds the keyword into a comparable volume identity.
TIFFLIST_FILENAME_PATTERN = re.compile(
    r'^(.*?)_?(?i:tifflist)\.(?i:txt|csv)$'
)
TOWNLIST_KEYWORD_PATTERN = re.compile(r'(?i:town[_ -]?list)')


def is_town_list_filename(filename):
    """True for any .xlsx/.csv file with the 'town list' keyword
    anywhere in its name (see TOWNLIST_KEYWORD_PATTERN)."""
    if not filename.lower().endswith((".xlsx", ".csv")):
        return False
    return bool(TOWNLIST_KEYWORD_PATTERN.search(filename))

# An existing merged ingest output file, e.g. "1885-1886-ingest.csv" -
# the two years are the earliest and latest issue dates actually
# CONTAINED in that file (they're the same year when a file happens to
# fit inside one calendar year). The optional "_partN" suffix is only
# ever used as a tie-breaker if two files would otherwise end up with
# the exact same year range.
INGEST_FILENAME_PATTERN = re.compile(
    r'^(\d{4})-(\d{4})-ingest(?:_part(\d+))?\.csv$', re.IGNORECASE
)


# =============================================================================
# STEP 1 - READ THE TOWN LIST FILE INTO A SIMPLE GRID
# =============================================================================
# Regardless of whether the Town List is .xlsx or .csv, we turn it into the
# same simple shape first: a list of rows, where each row is a list of
# plain-text cell values. Everything after this point works off that grid,
# so the rest of the code doesn't need to care which file format we started
# from.
# =============================================================================

def read_grid(source):
    """
    Read a Town List from EITHER:
        - a local file path ending in .xlsx or .csv, OR
        - a Google Sheets URL (the sheet just needs to be shared as
          "Anyone with the link can view" - no Google login needed)

    Either way, this returns the same simple list-of-lists "grid" shape,
    so nothing else in the script needs to know or care which kind of
    source the Town List actually came from.
    """
    if is_google_sheets_url(source):
        return fetch_google_sheet_grid(source)

    # Not a Google Sheets URL - treat it as a path to a local file.
    extension = os.path.splitext(source)[1].lower()

    if extension == ".xlsx":
        # openpyxl reads real Excel files. We only import it here (instead
        # of at the top of the file) so that people whose Town List is a
        # plain .csv (or a Google Sheet) never need to install openpyxl.
        import openpyxl

        workbook = openpyxl.load_workbook(source, data_only=True)
        worksheet = workbook.active  # the first/only sheet

        grid = []
        for row in worksheet.iter_rows(values_only=True):
            text_row = [cell_to_text(cell) for cell in row]
            grid.append(text_row)
        return grid

    elif extension == ".csv":
        with open(source, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            grid = [[cell.strip() for cell in row] for row in reader]
        return grid

    else:
        raise ValueError(
            f"Don't know how to read a Town List with extension "
            f"'{extension}'. Expected .xlsx, .csv, or a Google Sheets URL."
        )


def is_google_sheets_url(text):
    """
    True if the given string looks like a Google Sheets URL rather than
    a local file path. We just check whether it contains the
    "docs.google.com/spreadsheets/d/..." pattern that every Sheets URL
    has, regardless of whatever comes after it (/edit, #gid=123, etc.).
    """
    return bool(GOOGLE_SHEETS_URL_PATTERN.search(text))


def fetch_google_sheet_grid(sheet_url):
    """
    Download a Google Sheet as CSV text and parse it into the same
    list-of-lists "grid" shape that read_grid() produces for local files.

    HOW THIS WORKS: Google Sheets has a built-in "export as CSV" link for
    any sheet that's shared with at least "Anyone with the link can
    view". We build that export link ourselves from the URL you gave us,
    then download it with Python's built-in urllib (no extra libraries,
    no API key, no login needed) - it behaves just like clicking
    File > Download > CSV in your browser.
    """
    # Pull the long ID out of the URL, e.g. the "1AbCdEfGh..." part of
    # https://docs.google.com/spreadsheets/d/1AbCdEfGh.../edit
    id_match = GOOGLE_SHEETS_URL_PATTERN.search(sheet_url)
    if not id_match:
        raise ValueError(
            f"This doesn't look like a Google Sheets URL: {sheet_url}"
        )
    sheet_id = id_match.group(1)

    # If the URL specifies a particular tab (gid=...), use that tab.
    # Otherwise default to gid=0, which is the FIRST tab in the sheet.
    gid_match = GOOGLE_SHEETS_GID_PATTERN.search(sheet_url)
    gid = gid_match.group(1) if gid_match else "0"

    export_url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=csv&gid={gid}"
    )

    try:
        with urllib.request.urlopen(export_url) as response:
            raw_bytes = response.read()
    except urllib.error.HTTPError as err:
        # This almost always means the sheet isn't shared publicly yet.
        raise RuntimeError(
            f"Couldn't download the Google Sheet (HTTP error {err.code}). "
            f"Make sure the sheet's sharing setting is set to 'Anyone "
            f"with the link can view' (click Share, top right of the "
            f"sheet, then change 'Restricted' to 'Anyone with the "
            f"link'), then try again."
        ) from err
    except urllib.error.URLError as err:
        raise RuntimeError(
            f"Couldn't reach Google to download the sheet ({err.reason}). "
            f"Check your internet connection and try again."
        ) from err

    # The downloaded bytes are plain CSV text - decode and parse them
    # exactly the same way we'd parse a downloaded .csv file. "utf-8-sig"
    # quietly strips a byte-order-mark if Google includes one.
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    grid = [[cell.strip() for cell in row] for row in reader]
    return grid


def cell_to_text(value):
    """
    Convert one raw Excel cell value into clean text.

    The tricky case this handles: Excel often stores whole numbers like
    "Paper No. 39" internally as the FLOAT 39.0, not the integer 39. If we
    just did str(39.0) we'd get the text "39.0" - and later code that
    checks "is this text just digits?" (paper_no_text.isdigit()) would
    say NO, because of the decimal point, and silently skip that whole
    issue. We fix that here, once, so nothing downstream has to worry
    about it: any float that has no fractional part (39.0, 40.0, ...)
    gets turned into a plain integer string ("39", "40", ...) instead.
    """
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


# =============================================================================
# STEP 2 - TURN THE GRID INTO {paper_no: [town1, town2, ...]}
# =============================================================================
# This is the part that has to understand the TWO possible layouts.
# We look at the grid and decide which layout it is, then parse accordingly.
# =============================================================================

def parse_town_list(grid):
    """
    Parse a Town List grid (see read_grid) into a dictionary that maps:
        paper number (int)  ->  list of town names (list of str)

    Auto-detects which of THREE possible layouts the grid is in, and
    calls the matching parser below:
        Format A - one row per issue, "No." column, towns across columns
        Format B - "Paper No." header row, towns stacked down columns
        Format C - "Slim" layout with pre-filled Subject/Geographic
                   columns instead of plain town names (see
                   parse_format_c_slim for details)
    """
    if not grid or not grid[0]:
        raise ValueError("Town List file appears to be empty.")

    header_row = grid[0]
    header_lower = [cell.strip().lower() for cell in header_row]
    first_cell = header_row[0].lower()

    if "subject" in header_lower and "geographic" in header_lower:
        # A "Subject" AND "Geographic" column together is a distinctive
        # fingerprint of the Slim format - neither Format A nor B ever
        # has both of these as column headers.
        return parse_format_c_slim(grid)
    elif first_cell.startswith("paper"):
        # e.g. header_row[0] == "Paper No." -> issue numbers run ACROSS
        # this row, and towns are stacked DOWN each column below it.
        return parse_format_b_transposed(grid)
    else:
        # Otherwise assume Format A: one issue per row, with a "No." (or
        # similar) column holding the paper number.
        return parse_format_a_rows(grid)


def parse_format_a_rows(grid):
    """
    Parse "one row per issue" layout:

        No. | Subject 1  | Subject 2 | ...
        39  | Deep River | Chester   | ...

    or the "Old Method" variant, which adds a reference-filename column
    BEFORE the paper number:

        Filelist.txt              | No. | Subject 1  | ...
        18850102_vXI_n39.pdf      | 39  | Deep River | ...

    We find whichever column is the paper-number column (its header is
    something like "No." or "Paper No."), and every column after that is
    treated as a town name (blank cells are just skipped).

    If there's ALSO a column before it whose header looks like a
    filename reference (contains "file"), each row's volume is read
    from that filename via parse_issue_prefix() - this matters because
    paper numbers reset every volume, so a Town List spanning more than
    one volume (like the "Old Method" example above, which covers both
    Vol. XI and Vol. XII) needs a per-row volume to avoid collisions.
    Otherwise every row's volume is left as None, to be resolved later
    from context (see resolve_town_list_volumes()).

    Returns dict[(volume_or_None, paper_no)] -> list of town names.
    """
    header_row = grid[0]

    # Find the column whose header looks like a paper-number column.
    # We accept things like "No.", "No", "Paper No.", case-insensitively.
    paper_no_col = None
    for col_index, header_text in enumerate(header_row):
        cleaned = header_text.strip().lower().rstrip(".")
        if cleaned in ("no", "paper no", "papernum", "paper number"):
            paper_no_col = col_index
            break

    if paper_no_col is None:
        raise ValueError(
            "Couldn't find a paper-number column (looked for a header "
            "like 'No.' or 'Paper No.') in this Town List file."
        )

    reference_col = None
    for col_index in range(paper_no_col):
        if "file" in header_row[col_index].strip().lower():
            reference_col = col_index
            break

    town_list = {}

    # Walk every data row (skip the header row itself).
    for row in grid[1:]:
        if paper_no_col >= len(row):
            continue  # row is too short to even have a paper number

        paper_no_text = row[paper_no_col].strip()
        if not paper_no_text.isdigit():
            continue  # blank trailing row, or something we can't parse

        paper_no = int(paper_no_text)

        volume = None
        if reference_col is not None and reference_col < len(row):
            parsed = parse_issue_prefix(row[reference_col])
            if parsed:
                volume, _ = parsed

        # Every column AFTER the paper-number column is a potential town.
        # Blank cells are simply skipped (that's how the padding works).
        towns = [
            cell.strip()
            for cell in row[paper_no_col + 1:]
            if cell.strip()
        ]

        town_list[(volume, paper_no)] = towns

    return town_list


def parse_format_b_transposed(grid):
    """
    Parse "new method" layout:

        Paper No. | 39         | 40         | ...
        Towns     | Deep River | Deep River | ...
                   | Chester    | Chester    | ...

    One issue per COLUMN. Row 0 holds the paper numbers themselves
    (starting at column 1; column 0 is just the "Paper No." label).
    Every row after that holds one more town for each column, reading
    top-to-bottom until the cells run out (blank = no more towns).

    This format has no per-row volume reference, so every entry's
    volume is left as None (to be resolved later from context - see
    resolve_town_list_volumes()). Returns dict[(None, paper_no)] ->
    list of town names.
    """
    header_row = grid[0]
    town_list = {}

    # Walk every column starting at index 1 (index 0 is the row label).
    for col_index in range(1, len(header_row)):
        paper_no_text = header_row[col_index].strip()
        if not paper_no_text.isdigit():
            continue  # not a real paper-number column, skip it

        paper_no = int(paper_no_text)

        # Collect every non-blank town for this column, from row 1 down
        # to the bottom of the grid.
        towns = []
        for row in grid[1:]:
            if col_index < len(row):
                town = row[col_index].strip()
                if town:
                    towns.append(town)

        town_list[(None, paper_no)] = towns

    return town_list


SUBJECT_ENTRY_PATTERN = re.compile(r'^Newspaper -- (.+), CT$')


def parse_format_c_slim(grid):
    """
    Parse the "Slim" layout some contributors have started sending,
    where the Subject and Geographic Subject columns are ALREADY
    filled in with Islandora's exact expected wording, rather than
    plain town names:

        Filelist.txt              | ... | Subject                          | Geographic
        19090115_vXXXVI_n01.pdf   | ... | Newspaper -- Deep River, CT^^...  | Deep River (Conn.)^^...

    Rather than trust the Geographic column on its own (which could in
    theory drift out of sync with Subject if someone hand-edits one and
    not the other), we pull the plain town names back OUT of the
    Subject column instead - stripping the "Newspaper -- " prefix and
    ", CT" suffix off each ^^-separated entry - and hand those back to
    build_subject_and_geo() later, exactly like every other format.
    This keeps the final CSV's formatting identical no matter which
    Town List format supplied it, and means a typo in the Geographic
    column alone can't slip through uncaught.

    Each issue is identified by parsing the reference filename in
    column 0 (the same pattern "Old Method" Format A files use for
    their reference column - see parse_issue_prefix), so this format
    supplies its own per-row volume; no fallback is needed.

    Any row whose filename doesn't parse, or whose Subject cell is
    blank, is skipped quietly - it'll simply show up later as a
    missing_town_list problem for that issue, same as any other gap.

    Returns dict[(volume, paper_no)] -> list of town names, same shape
    as every other format's parser.
    """
    header_row = grid[0]
    header_lower = [cell.strip().lower() for cell in header_row]
    subject_col = header_lower.index("subject")

    town_list = {}

    for row in grid[1:]:
        if not row or not row[0].strip():
            continue

        parsed = parse_issue_prefix(row[0])
        if parsed is None:
            continue  # not a recognizable "yyyymmdd_vVOL_nNN..." filename

        volume, paper_no = parsed

        towns = []
        if subject_col < len(row) and row[subject_col].strip():
            for entry in row[subject_col].split(MULTI_SEP):
                match = SUBJECT_ENTRY_PATTERN.match(entry.strip())
                if match:
                    towns.append(match.group(1))

        town_list[(volume, paper_no)] = towns

    return town_list


# =============================================================================
# STEP 3 - GET TIFF FILENAMES AND GROUP PAGES INTO ISSUES
# =============================================================================
# There are two ways to tell this script about your TIFFs:
#   (a) point it at a real folder full of TIFFs on your computer, or
#   (b) point it at a CSV/text file that just LISTS the filenames (handy
#       when the actual TIFFs live somewhere you can't browse directly,
#       like a remote ingest server).
# Both paths end up producing the exact same "issues_by_paper_no" shape,
# via the shared parse_tiff_filename() helper below.
# =============================================================================

def int_to_roman(n):
    """Convert a plain integer (1, 2, 3...) to upper-case roman numerals
    ("I", "II", "III"...). Newspaper volume numbers are small, so this
    doesn't need to handle huge numbers."""
    value_symbols = [
        (1000, "M"), (900, "CM"), (500, "D"), (400, "CD"),
        (100, "C"), (90, "XC"), (50, "L"), (40, "XL"),
        (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I"),
    ]
    result = []
    for value, symbol in value_symbols:
        while n >= value:
            result.append(symbol)
            n -= value
    return "".join(result)


def normalize_volume(volume_text):
    """
    Volumes are normally written as roman numerals ("XI"), but some
    early files use a plain number instead ("1", "11"). This makes sure
    both spellings of the SAME volume come out identical, so e.g.
    "v1" and "vI" don't get treated as two different volumes.
    """
    if volume_text.isdigit():
        return int_to_roman(int(volume_text))
    return volume_text.upper()


def extract_volume_token(prefix):
    """
    Given whatever text came before the "tifflist" / "Town List" keyword
    in an input filename (e.g. "Vol_09_", "vI_", "09-", "" for a bare
    "tifflist.txt"), strip off a "Vol"/"v" label prefix if there is one,
    leaving just the volume identifier itself for normalize_volume() to
    compare. If there's no recognizable "Vol"/"v" prefix, the whole
    (trimmed) prefix is used as-is - so a plain "09_tifflist.txt" still
    works even without a "Vol_" label.
    """
    prefix = prefix.strip(" _-")
    if not prefix:
        return ""

    match = re.match(r'(?i)^vol[_ -]?', prefix)
    if match:
        return prefix[match.end():].strip(" _-")

    # A bare lowercase "v" followed by all-uppercase letters is treated
    # as a roman-numeral label (e.g. "vI", "vIX") - checking case avoids
    # misreading a roman numeral that itself starts with "V" (e.g. "VI"
    # alone, with no separate "v" label) as having a redundant prefix.
    if prefix[0] == "v" and len(prefix) > 1 and prefix[1:].isalpha() and prefix[1:] == prefix[1:].upper():
        return prefix[1:]

    return prefix


ISSUE_PREFIX_PATTERN = re.compile(r'^(\d{8})_v([A-Za-z0-9]+)_n(\d+)')


def parse_issue_prefix(text):
    """
    Pull (volume, paper_no) out of the START of a reference string like
    "18850102_vXI_n39.pdf" or "18850102_vXI_n39" - used for "Old Method"
    Town List files, whose first column names the issue by filename
    (with no page number, and often a .pdf extension) rather than a
    bare paper number. Returns None if it doesn't look like one of these.
    """
    match = ISSUE_PREFIX_PATTERN.match(text.strip())
    if not match:
        return None
    _, volume_raw, paper_no_text = match.groups()
    return normalize_volume(volume_raw), int(paper_no_text)


def volume_from_town_list_filename(filename):
    """
    Derive a volume label from a Town List's OWN filename (e.g.
    "Vol_1_Town_List.xlsx" -> "I"), for the common case where the whole
    file is just one volume's worth of plain paper numbers with no
    per-row volume reference. Returns None if no volume label can be
    found in the filename at all.
    """
    name = os.path.splitext(filename)[0]
    prefix = TOWNLIST_KEYWORD_PATTERN.sub("", name)
    token = extract_volume_token(prefix)
    return normalize_volume(token) if token else None


def parse_tiff_filename(filename):
    """
    Pull the useful pieces out of one TIFF filename, e.g.
    "18850116_vXI_n41_0001.tif" ->

        issue_id   = "18850116_vXI_n41"   (filename with page # removed)
        date_iso   = "1885-01-16"
        volume     = "XI"
        paper_no   = 41                   (as an int)
        page_number = 1                   (as an int, defaults to 1 if
                                            the filename has no page #)

    Returns None if the filename doesn't match TIFF_NAME_PATTERN at all
    (callers are expected to check for that and handle it themselves).
    """
    match = TIFF_NAME_PATTERN.match(filename)
    if not match:
        return None

    date_text, volume_raw, paper_no_text, page_text = match.groups()
    volume = normalize_volume(volume_raw)
    paper_no = int(paper_no_text)

    # Turn "18850116" into a readable date like "1885-01-16".
    # If a filename ever has a nonsense date, we don't want the whole
    # script to crash - fall back to the raw text instead.
    try:
        date_obj = datetime.strptime(date_text, "%Y%m%d")
        date_iso = date_obj.strftime("%Y-%m-%d")
    except ValueError:
        date_iso = date_text  # fall back to raw digits

    issue_id = f"{date_text}_v{volume}_n{paper_no_text}"

    # The page number group is OPTIONAL in the filename pattern (see
    # TIFF_NAME_PATTERN above). If a filename has no page number at all,
    # we just use 1 - it's almost certainly a single-page issue.
    page_number = int(page_text) if page_text else 1

    return issue_id, date_iso, volume, paper_no, page_number


def scan_tiff_directory(tiff_dir):
    """
    Look at every .tif/.tiff file in a real folder on disk, parse each
    filename, and group the individual PAGE files into ISSUES.

    Returns a dictionary that maps:
        (volume, paper number) -> {
            "issue_id": "18850116_vXI_n41",   # filename with page # removed
            "date":     "1885-01-16",         # ISO-formatted date string
            "volume":   "XI",
            "pages": [                        # one entry per page TIFF,
                {"filename": "...0001.tif", "page_number": 1,
                 "digital_file": "/opt/ingest_data/uploads/3/tiff/...0001.tif"},
                ...
            ]
        }

    Keying by (volume, paper number) together - rather than just paper
    number - matters because paper numbering resets every volume, so
    two different volumes can both have a "paper #1"; keying by paper
    number alone would silently merge them.

    Files that don't match the expected naming pattern are collected
    separately and returned too, so the calling code can report them.
    """
    issues_by_key = {}
    unrecognized_files = []

    for filename in sorted(os.listdir(tiff_dir)):
        # Only look at .tif / .tiff files - skip everything else
        # (this also naturally skips subfolders, .DS_Store, etc.)
        if not filename.lower().endswith((".tif", ".tiff")):
            continue

        parsed = parse_tiff_filename(filename)
        if parsed is None:
            unrecognized_files.append(filename)
            continue

        issue_id, date_iso, volume, paper_no, page_number = parsed
        key = (volume, paper_no)

        # First time we've seen this volume+paper number? Start a new entry.
        if key not in issues_by_key:
            issues_by_key[key] = {
                "issue_id": issue_id,
                "date": date_iso,
                "volume": volume,
                "pages": [],
            }

        issues_by_key[key]["pages"].append({
            "filename": filename,
            "page_number": page_number,
            # We're scanning a real local folder here, so we don't know
            # the eventual server-side path - build it ourselves using
            # the issue's year, per the standard server layout.
            "digital_file": build_digital_file(date_iso, filename),
        })

    return issues_by_key, unrecognized_files


def read_tiff_list_file(list_file_path):
    """
    Read TIFF filenames out of a plain CSV or text "listing" file instead
    of scanning an actual folder on disk. Handles TWO kinds of listing
    files automatically - you don't need to tell us which one you have:

        FORMAT 1 - one full path per line, e.g.:
            /opt/ingest_data/uploads/3/1885/18850102_vXI_n39_0001.tif

        FORMAT 2 - a Windows "dir" command listing, e.g.:
            06/24/2025  08:05 AM     9,082,646  18850612_vXII_n10_0005.tif

    For every line, we just look for a TIFF filename matching our naming
    pattern ANYWHERE in the line (see TIFF_NAME_SEARCH_PATTERN) and pull
    it out - everything else on the line (a directory path, a date/time/
    size) is ignored. The "digital_file" value is always BUILT FRESH
    ourselves from the issue's year (see build_digital_file), even if the
    line already contained a full path - some older listings show a
    volume-based folder instead of a year-based one, and the server has
    standardized on year-based folders going forward, so we don't trust
    whatever path text happens to already be in the file.

    Returns the exact same shape as scan_tiff_directory(): a dictionary
    of issues keyed by (volume, paper number), plus a list of any lines
    we couldn't make sense of.
    """
    issues_by_key = {}
    unrecognized_lines = []

    # "utf-8-sig" quietly strips a byte-order-mark if the file has one
    # (common in files saved from Windows tools). Universal newline
    # handling (the default in text mode) takes care of \r\n line
    # endings automatically, so we don't need to worry about those.
    with open(list_file_path, "r", encoding="utf-8-sig") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue  # skip blank lines

            match = TIFF_NAME_SEARCH_PATTERN.search(line)
            if not match:
                unrecognized_lines.append(line)
                continue

            filename = match.group(1)

            parsed = parse_tiff_filename(filename)
            if parsed is None:
                # Shouldn't normally happen (the search pattern is a
                # looser version of the same rule), but just in case:
                unrecognized_lines.append(line)
                continue

            issue_id, date_iso, volume, paper_no, page_number = parsed
            key = (volume, paper_no)

            if key not in issues_by_key:
                issues_by_key[key] = {
                    "issue_id": issue_id,
                    "date": date_iso,
                    "volume": volume,
                    "pages": [],
                }

            issues_by_key[key]["pages"].append({
                "filename": filename,
                "page_number": page_number,
                "digital_file": build_digital_file(date_iso, filename),
            })

    return issues_by_key, unrecognized_lines


# =============================================================================
# STEP 4 - TURN A TOWN LIST INTO subject / geographic_subject TEXT
# =============================================================================

def build_subject_and_geo(towns):
    """
    Turn a list of town names into the two formatted strings that go into
    the "subject" and "geographic_subject" columns.

        towns = ["Chester", "Essex"]

        subject_val            -> "Newspaper -- Chester, CT^^Newspaper -- Essex, CT"
        geographic_subject_val -> "Chester (Conn.)^^Essex (Conn.)"

    If towns is an empty list (no Town List match for this issue), both
    returned strings are just "" - left blank for a volunteer to fill in.
    """
    subject_val = MULTI_SEP.join(f"Newspaper -- {town}, CT" for town in towns)
    geographic_subject_val = MULTI_SEP.join(f"{town} (Conn.)" for town in towns)
    return subject_val, geographic_subject_val


def format_date_long(date_iso):
    """
    Turn an ISO date string ("1885-01-16") into the long display format
    used in titles ("Jan 16, 1885") - month abbreviation, day WITHOUT a
    leading zero, four-digit year.
    """
    date_obj = datetime.strptime(date_iso, "%Y-%m-%d")
    # strftime's %-d (no leading zero) isn't reliably cross-platform, so
    # we build the day number ourselves with plain str() instead.
    return f"{date_obj.strftime('%b')} {date_obj.day}, {date_obj.year}"


# =============================================================================
# STEP 5 - BUILD THE FULL, 64-COLUMN INGEST ROWS
# =============================================================================

def compute_issue_record(key, issue, town_lists):
    """
    Gather everything needed to describe ONE issue - title, subject/geo,
    origin_information, etc - WITHOUT assigning any ID numbers yet. This
    is the "content" of an issue, independent of where it ends up in a
    CSV, which is what lets the same issue be safely compared against an
    already-written row later (for by-year merging/dedup).

    key - (volume, paper_no), matching how tiff_issues and town_lists
          are both keyed (see scan_tiff_directory / read_tiff_list_file
          and resolve_town_list_volumes)
    """
    volume, paper_no = key
    towns = town_lists.get(key, [])
    subject_val, geo_val = build_subject_and_geo(towns)

    issue_num_padded = f"{paper_no:02d}"
    date_long = format_date_long(issue["date"])
    title = (
        f"{NEWSPAPER_NAME} - {date_long}; "
        f"Vol. {issue['volume']} No. {issue_num_padded}"
    )

    return {
        "issue_id": issue["issue_id"],
        "date": issue["date"],
        "year": issue["date"][:4],
        "volume": issue["volume"],
        "paper_no": paper_no,
        "pages": sorted(issue["pages"], key=lambda p: p["page_number"]),
        "title": title,
        "subject": subject_val,
        "geographic_subject": geo_val,
        "persons": build_persons(issue["date"]),
        "origin_information": build_origin_information(issue["date"], NEWSPAPER_NAME),
    }


def build_parent_row_dict(record):
    """
    Build the "Publication Issue" parent row's column values for one
    issue record, WITHOUT setting ID (that's assigned separately by
    rows_from_issue_record, since it depends on where this issue lands
    in a CSV that may already have other rows in it).

    local_identifier is set to the issue_id (e.g. "18850116_vXI_n41") -
    this is what lets by-year merging recognize "this issue is already
    in the file" on a later run.
    """
    row = {column: "" for column in COLUMNS}
    row.update({
        "member_of_existing_entity_id": MEMBER_OF_EXISTING_ENTITY_ID,
        "publish": PUBLISH,
        "model": "Publication Issue",
        "held_by": HELD_BY,
        "title": record["title"],
        "rights_statement": RIGHTS_STATEMENT,
        "resource_type": RESOURCE_TYPE,
        "description": DESCRIPTION,
        "persons": record["persons"],
        "origin_information": record["origin_information"],
        "genre": GENRE,
        "subject": record["subject"],
        "geographic_subject": record["geographic_subject"],
        "notes": load_notes(),
        "local_identifier": record["issue_id"],
    })
    return row


def rows_from_issue_record(record, start_id):
    """
    Turn one issue record into an actual list of CSV rows (one parent
    "Publication Issue" row, then one "Page" row per TIFF page),
    assigning ID numbers starting at start_id.

    Returns (rows, next_id) where next_id is whatever ID the NEXT issue
    should start at.
    """
    next_id = start_id
    parent_id = next_id
    next_id += 1

    parent_row = build_parent_row_dict(record)
    parent_row["ID"] = parent_id
    rows = [parent_row]

    for page in record["pages"]:
        page_id = next_id
        next_id += 1

        page_row = {column: "" for column in COLUMNS}
        page_row.update({
            "ID": page_id,
            "member_of": parent_id,   # links this page back to its issue
            "publish": PUBLISH,
            "model": "Page",
            "held_by": HELD_BY,
            "title": f"{record['title']} Page {page['page_number']}",
            "rights_statement": RIGHTS_STATEMENT,
            "weight": page["page_number"],
            "digital_file": page["digital_file"],
            "media_use": "Original File",
            "digital_origin": "reformatted digital",
            "resource_type": RESOURCE_TYPE,
        })
        rows.append(page_row)

    return rows, next_id


def chunk_records_by_line_limit(records, start_line_count=1,
                                 target_lines=TARGET_LINES_PER_FILE):
    """
    Split an ORDERED list of issue records into chunks sized so that no
    chunk's row count (header + rows) intentionally exceeds
    target_lines - except that a single issue's rows are NEVER split
    across two chunks, so one unusually large issue can push a chunk
    over the target by itself.

    A chunk boundary is ALSO forced whenever there's a gap of more than
    MAX_CONTIGUOUS_YEAR_GAP years between one issue and the next (e.g.
    an 1876 issue followed by an 1885 issue) - even if there's still
    plenty of room left under target_lines - so that a large historical
    gap doesn't get lumped into one oddly-named file like
    "1876-1885-ingest.csv".

    start_line_count lets the FIRST chunk account for a header + rows
    that already exist in a file being appended to (pass 1 for a brand
    new, empty file). Every chunk after the first always starts fresh
    at a new file's header line (1).

    Returns a list of chunks, where each chunk is a list of records.
    """
    chunks = []
    current_chunk = []
    current_lines = start_line_count
    current_chunk_end_year = None

    for record in records:
        issue_line_count = 1 + len(record["pages"])  # parent row + pages
        year_gap = (
            current_chunk_end_year is not None
            and int(record["year"]) - int(current_chunk_end_year) > MAX_CONTIGUOUS_YEAR_GAP
        )
        if current_chunk and (current_lines + issue_line_count > target_lines or year_gap):
            chunks.append(current_chunk)
            current_chunk = []
            current_lines = 1  # a new file, just the header so far

        current_chunk.append(record)
        current_lines += issue_line_count
        current_chunk_end_year = record["year"]

    if current_chunk:
        chunks.append(current_chunk)

    return chunks




# =============================================================================
# STEP 6 - WRITE BOTH OUTPUT FILES
# =============================================================================

# =============================================================================
# INPUT SCANNING - find Vol_xx_tifflist.txt / Vol_xx_Town_List.xlsx pairs
# in a directory (used by --by-year / merge mode below).
# =============================================================================

def file_contains_tiff_filenames(path):
    """
    Peek at a file's CONTENT (not its name) and decide whether it looks
    like a tifflist - i.e. whether most of its non-blank lines contain
    a recognizable TIFF filename (see TIFF_NAME_SEARCH_PATTERN).

    This is a fallback for files whose name doesn't cleanly follow the
    "...tifflist.txt" convention - e.g. a contributor naming their file
    "tifflist07282026.txt" (a date tacked on after the keyword, so
    TIFFLIST_FILENAME_PATTERN doesn't match it) - so a file isn't
    skipped just because of how it happened to get named.

    Requiring a strong MAJORITY of lines to match (not just one) avoids
    misidentifying some unrelated file that merely mentions a TIFF
    filename once, e.g. in a comment or a report.
    """
    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            lines = [line for line in f if line.strip()]
    except (UnicodeDecodeError, OSError):
        return False  # not a readable text file at all

    if not lines:
        return False

    matching = sum(1 for line in lines if TIFF_NAME_SEARCH_PATTERN.search(line))
    return matching / len(lines) >= 0.5


def find_tiff_source_files(directory):
    """
    Every file in `directory` that's a tifflist - matched primarily by
    NAME (see TIFFLIST_FILENAME_PATTERN), with a CONTENT-based fallback
    (see file_contains_tiff_filenames) for .txt/.csv files that don't
    follow the naming convention but are clearly full of TIFF filenames
    once you actually look inside.

    Files already recognized as a Town List by name are never
    considered here, even if they happened to contain TIFF-shaped text
    somewhere - a file can't be both.
    """
    by_name = []
    maybe_by_content = []

    for f in sorted(os.listdir(directory)):
        path = os.path.join(directory, f)
        if not os.path.isfile(path):
            continue
        if TIFFLIST_FILENAME_PATTERN.match(f):
            by_name.append(f)
        elif f.lower().endswith((".txt", ".csv")) and not is_town_list_filename(f):
            maybe_by_content.append(f)

    by_content = [
        f for f in maybe_by_content
        if file_contains_tiff_filenames(os.path.join(directory, f))
    ]
    for f in by_content:
        print(f"  (recognized '{f}' as a tifflist by its content, not its name)")

    return sorted(by_name + by_content)


def find_town_list_source_files(directory):
    """Every file in `directory` that looks like a Town List, by name."""
    return sorted(f for f in os.listdir(directory) if is_town_list_filename(f))


def resolve_town_list_volumes(town_lists, tiff_issues, filename_hint=None):
    """
    town_lists may contain entries keyed (None, paper_no) - meaning "we
    don't know which volume this row belongs to" (a plain "No." column
    with no per-row filename reference). This resolves those using
    whatever context is available:

        1. Try to read a volume label off the Town List's OWN filename
           (e.g. "Vol_1_Town_List.xlsx" -> "I") - this is the normal,
           reliable case, since standard-format Town Lists are named
           per volume by convention.
        2. If the filename has no discernible volume label at all (e.g.
           a bare "Town_List.csv"), and every TIFF issue read so far
           happens to belong to a single volume, assume the Town List's
           rows belong to that same volume.
        3. If neither works, those rows can't be matched to anything -
           they're dropped, and a problem string is returned explaining
           why.

    Returns (resolved_town_lists, problems) - resolved_town_lists has
    every key as (volume, paper_no), never (None, paper_no).
    """
    none_keys = [k for k in town_lists if k[0] is None]
    if not none_keys:
        return town_lists, []

    resolved = {k: v for k, v in town_lists.items() if k[0] is not None}
    problems = []

    fallback_volume = volume_from_town_list_filename(filename_hint) if filename_hint else None
    if fallback_volume is None:
        candidate_volumes = {k[0] for k in tiff_issues}
        if len(candidate_volumes) == 1:
            fallback_volume = next(iter(candidate_volumes))

    if fallback_volume:
        for (_, paper_no) in none_keys:
            resolved[(fallback_volume, paper_no)] = town_lists[(None, paper_no)]
    else:
        problems.append(
            f"[unresolved_volume] {len(none_keys)} row(s) use plain paper "
            f"numbers with no way to tell which volume they belong to "
            f"(no per-row filename reference, no single matching volume "
            f"among the TIFF data, and no volume label in this file's own "
            f"name) - these rows were skipped."
        )

    return resolved, problems


def gather_all_records(directory):
    """
    Read EVERY recognized tifflist file and EVERY recognized Town List
    file in `directory` - regardless of how they're individually named,
    or whether they line up 1:1 - and combine everything into one pool
    of issue records, joined by (volume, paper number) rather than by
    which specific files they happened to come from. This is what lets
    a single tifflist file spanning multiple volumes (e.g. "tifflist.txt"
    covering both Vol. XI and Vol. XII) match correctly against a Town
    List using either the plain "No." format or the "Old Method" format
    (a per-row filename reference column).

    Returns (records, problems, tiff_filenames, townlist_filenames).
    """
    tiff_filenames = find_tiff_source_files(directory)
    townlist_filenames = find_town_list_source_files(directory)

    all_problems = []

    # ---- Read every tifflist file into one combined pool ---------------
    all_tiff_issues = {}
    for filename in tiff_filenames:
        path = os.path.join(directory, filename)
        print(f"Reading tifflist: {filename}")
        issues, unrecognized = read_tiff_list_file(path)

        for key, issue in issues.items():
            if key in all_tiff_issues:
                all_problems.append(
                    f"[{filename}] [duplicate_tiff_entry] Vol_{key[0]} "
                    f"paper #{key[1]} was already read from another "
                    f"tifflist file - keeping the first one seen."
                )
                continue
            all_tiff_issues[key] = issue

        for item in unrecognized:
            all_problems.append(
                f"[{filename}] [unrecognized_filename] '{item}' does not "
                f"match the expected pattern and was skipped."
            )

    # ---- Read every Town List file into one combined pool ---------------
    all_town_lists = {}
    for filename in townlist_filenames:
        path = os.path.join(directory, filename)
        print(f"Reading Town List: {filename}")
        grid = read_grid(path)
        raw_town_list = parse_town_list(grid)
        resolved, resolve_problems = resolve_town_list_volumes(
            raw_town_list, all_tiff_issues, filename_hint=filename
        )
        all_problems.extend(f"[{filename}] {p}" for p in resolve_problems)

        for key, towns in resolved.items():
            if key in all_town_lists:
                all_problems.append(
                    f"[{filename}] [duplicate_town_list_entry] "
                    f"Vol_{key[0]} paper #{key[1]} was already read from "
                    f"another Town List file - keeping the first one seen."
                )
                continue
            all_town_lists[key] = towns

    source_label = (
        f"{len(tiff_filenames)} tifflist file(s)" if tiff_filenames
        else "no tifflist files found"
    )
    records, join_problems = records_from_single_source(
        all_tiff_issues, all_town_lists, [], source_label=source_label,
    )
    all_problems.extend(join_problems)

    return records, all_problems, tiff_filenames, townlist_filenames


# =============================================================================
# MERGE MODE (--by-year) - read ALL Vol_xx pairs in a directory at once,
# and combine every issue into a small number of "start-end-ingest.csv"
# files, named after the actual earliest/latest issue date each file
# contains (the two years are the same when a file happens to fit
# entirely inside one calendar year).
#
# Safe to re-run (e.g. on a schedule via GitHub Actions): every issue is
# checked against what's already on disk (by issue_id, stored in
# local_identifier) before anything is added, the same way a human
# reviewer would:
#   - not found anywhere -> append as a new issue (to the most recent
#     file if it still has room under TARGET_LINES_PER_FILE, otherwise
#     start a new file)
#   - found, and matches  -> already there, skip quietly
#   - found, but differs  -> flag for review, don't overwrite
#
# Appending to the most recent file can extend its date range, in which
# case the file is renamed to match (e.g. "1885-1885-ingest.csv" growing
# to "1885-1886-ingest.csv"). A file that's already at/over the line
# target is left alone and a new file is started instead.
# =============================================================================

def write_matching_xlsx(csv_path):
    """
    Write an .xlsx copy of an ingest CSV file that was just written or
    updated, so contributors who prefer opening things in Excel don't
    have to convert it by hand.

    This is purely a courtesy copy - the CSV remains the file this
    script actually reads back in on future runs (see
    find_ingest_files / INGEST_FILENAME_PATTERN, which only ever
    matches ".csv"). The .xlsx is regenerated fresh from the CSV's
    final content every time it changes, and is never itself read by
    anything - so there's no risk of the two ever silently disagreeing.
    """
    import openpyxl

    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"
    workbook.save(xlsx_path)


def normalize_for_compare(row):
    """
    Prepare a parent row dict for an equality check against another copy
    of "the same issue", ignoring the two things that are EXPECTED to
    differ between runs even when nothing meaningful changed: the
    ID/member_of numbers (they depend on file position, not content),
    and the date_captured sub-field inside origin_information (it's
    always "today", so it changes every single run by definition).
    """
    normalized = {
        k: str(v) for k, v in row.items() if k not in ("ID", "member_of")
    }
    parts = normalized.get("origin_information", "").split("|")
    if len(parts) == 13:
        parts[4] = ""  # blank out date_captured before comparing
        normalized["origin_information"] = "|".join(parts)
    return normalized


def find_ingest_files(directory):
    """
    Scan `directory` for existing "start-end-ingest.csv" files and read
    each one's rows. Returns a list of dicts (one per file), each with
    filename, path, start_year, end_year, part_num, and rows - sorted so
    the CHRONOLOGICALLY LATEST file (by end_year, then start_year, then
    part_num) comes last.
    """
    files = []
    for filename in os.listdir(directory):
        match = INGEST_FILENAME_PATTERN.match(filename)
        if not match:
            continue
        start_year, end_year, part_text = match.groups()
        path = os.path.join(directory, filename)
        with open(path, "r", newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        files.append({
            "filename": filename,
            "path": path,
            "start_year": start_year,
            "end_year": end_year,
            "part_num": int(part_text) if part_text else 1,
            "rows": rows,
        })

    files.sort(key=lambda info: (info["end_year"], info["start_year"], info["part_num"]))
    return files


def make_ingest_filename(directory, start_year, end_year, taken_names):
    """
    Build a "start-end-ingest.csv" filename, adding a "_partN" suffix
    ONLY if that exact name is already taken (either already on disk, or
    already used earlier in this same run) - this should be rare, since
    files are chronological and rarely land on the exact same range.
    """
    candidate = f"{start_year}-{end_year}-ingest.csv"
    part_num = 2
    while candidate in taken_names or os.path.exists(os.path.join(directory, candidate)):
        candidate = f"{start_year}-{end_year}-ingest_part{part_num}.csv"
        part_num += 1
    return candidate


def records_from_single_source(tiff_issues, town_lists, unrecognized_items,
                                source_label, problem_prefix=""):
    """
    Turn one already-read (tiff_issues, town_lists) pair into a list of
    issue records plus a list of human-readable problem descriptions
    (missing_tiffs / missing_town_list / unrecognized_filename) - shared
    by both the default directory-gathering mode and manual single-pair
    mode.

    tiff_issues / town_lists are both keyed by (volume, paper_no) - see
    scan_tiff_directory / read_tiff_list_file and
    resolve_town_list_volumes.

    problem_prefix is prepended to every problem line (e.g. "[Vol_09] ")
    so problems from multiple sources can be told apart in one report;
    leave it blank when there's only a single source.
    """
    records = []
    problems = []

    town_keys = set(town_lists.keys())
    tiff_keys = set(tiff_issues.keys())
    only_in_town_list = town_keys - tiff_keys
    only_in_tiffs = tiff_keys - town_keys

    for key in sorted(only_in_town_list):
        volume, paper_no = key
        problems.append(
            f"{problem_prefix}[missing_tiffs] Town List has Vol_{volume} "
            f"paper #{paper_no} with towns "
            f"({MULTI_SEP.join(town_lists[key])}), but no matching TIFF "
            f"was found in {source_label}."
        )

    for item in unrecognized_items:
        problems.append(
            f"{problem_prefix}[unrecognized_filename] '{item}' does not "
            f"match the expected pattern and was skipped."
        )

    for key, issue in tiff_issues.items():
        record = compute_issue_record(key, issue, town_lists)
        record["source_volume"] = f"Vol_{key[0]}"
        if key in only_in_tiffs:
            problems.append(
                f"{problem_prefix}[missing_town_list] Issue "
                f"{record['issue_id']} has no Town List entry - included "
                f"with blank subject/geographic_subject."
            )
        records.append(record)

    return records, problems


def merge_and_write_records(directory, all_records, volume_problems, source_line):
    """
    The shared merge/dedup/write engine, used by BOTH the default
    directory-scanning mode and manual single-pair mode. See the module
    comment above run_by_year() for the merge rules and file-naming
    convention.

    directory        - where existing "start-end-ingest.csv" files live,
                        and where new/updated ones get written
    all_records      - every issue record to merge in (from
                        compute_issue_record, with "source_volume" set)
    volume_problems  - problem description strings (missing_tiffs etc.)
    source_line       - one line of text for the report describing where
                        all_records came from
    """
    # ---- Load whatever ingest files already exist -----------------------
    existing_files = find_ingest_files(directory)
    existing_by_id = {}
    for info in existing_files:
        for row in info["rows"]:
            if row.get("model") == "Publication Issue" and row.get("local_identifier"):
                existing_by_id[row["local_identifier"]] = row

    # ---- Sort out matched / mismatched / genuinely-new issues -----------
    # counts_by_source tracks these same three outcomes PER source_volume
    # (e.g. per tifflist file), so the report can show a breakdown rather
    # than just one combined total - otherwise there's no way to confirm
    # from the report alone that every source's issues actually made it in.
    new_records = []
    matched_count = 0
    mismatch_lines = []
    counts_by_source = {}

    def bump(source_volume, outcome, page_count):
        entry = counts_by_source.setdefault(
            source_volume,
            {"new": 0, "matched": 0, "mismatch": 0,
             "new_pages": 0, "matched_pages": 0, "mismatch_pages": 0},
        )
        entry[outcome] += 1
        entry[f"{outcome}_pages"] += page_count

    for record in sorted(all_records, key=lambda r: r["date"]):
        issue_id = record["issue_id"]
        page_count = len(record["pages"])

        if issue_id in existing_by_id:
            new_parent = normalize_for_compare(build_parent_row_dict(record))
            old_parent = normalize_for_compare(existing_by_id[issue_id])
            if new_parent == old_parent:
                matched_count += 1
                bump(record["source_volume"], "matched", page_count)
            else:
                mismatch_lines.append(
                    f"  - {issue_id} ({record['source_volume']}) already "
                    f"exists but the new data differs. NOT overwritten - "
                    f"please review manually."
                )
                bump(record["source_volume"], "mismatch", page_count)
            continue

        new_records.append(record)
        bump(record["source_volume"], "new", page_count)

    new_pages_total = sum(c["new_pages"] for c in counts_by_source.values())
    matched_pages_total = sum(c["matched_pages"] for c in counts_by_source.values())
    mismatch_pages_total = sum(c["mismatch_pages"] for c in counts_by_source.values())

    # ---- Write the new issues: continue the most recent file if it has
    # room, otherwise start fresh; roll into further new files as needed.
    written_files = []
    added_total = 0

    if new_records:
        active_file = None
        if existing_files:
            candidate = existing_files[-1]  # chronologically latest
            new_records_min_year = min(int(r["year"]) for r in new_records)
            year_gap = new_records_min_year - int(candidate["end_year"]) > MAX_CONTIGUOUS_YEAR_GAP
            if 1 + len(candidate["rows"]) < TARGET_LINES_PER_FILE and not year_gap:
                active_file = candidate

        taken_names = {info["filename"] for info in existing_files}
        start_line_count = 1 + len(active_file["rows"]) if active_file else 1
        chunks = chunk_records_by_line_limit(
            new_records, start_line_count=start_line_count
        )

        for i, chunk in enumerate(chunks):
            chunk_rows_out = []

            if i == 0 and active_file is not None:
                # Continue the existing active file, in place.
                existing_rows = active_file["rows"]
                next_id = (
                    max(int(r["ID"]) for r in existing_rows if r.get("ID")) + 1
                    if existing_rows else 1
                )
                for record in chunk:
                    rows, next_id = rows_from_issue_record(record, next_id)
                    chunk_rows_out.extend(rows)

                with open(active_file["path"], "a", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=COLUMNS)
                    writer.writerows(chunk_rows_out)

                # The file's content range may have grown - rename it to
                # match if so.
                new_start = min(active_file["start_year"], *(r["year"] for r in chunk))
                new_end = max(active_file["end_year"], *(r["year"] for r in chunk))
                final_path = active_file["path"]
                if (new_start, new_end) != (active_file["start_year"], active_file["end_year"]):
                    taken_names.discard(active_file["filename"])
                    new_filename = make_ingest_filename(directory, new_start, new_end, taken_names)
                    new_path = os.path.join(directory, new_filename)
                    os.rename(active_file["path"], new_path)
                    final_path = new_path
                    taken_names.add(new_filename)
                else:
                    taken_names.add(active_file["filename"])

                written_files.append(final_path)

            else:
                # A brand-new file, named after this chunk's own date range.
                next_id = 1
                for record in chunk:
                    rows, next_id = rows_from_issue_record(record, next_id)
                    chunk_rows_out.extend(rows)

                start_year = min(r["year"] for r in chunk)
                end_year = max(r["year"] for r in chunk)
                filename = make_ingest_filename(directory, start_year, end_year, taken_names)
                path = os.path.join(directory, filename)

                with open(path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=COLUMNS)
                    writer.writeheader()
                    writer.writerows(chunk_rows_out)

                taken_names.add(filename)
                written_files.append(path)

            added_total += len(chunk)

    # Every ingest CSV that was written or updated this run gets a
    # matching .xlsx copy, regenerated fresh from the CSV's final
    # content (see write_matching_xlsx).
    for path in written_files:
        write_matching_xlsx(path)

    # ---- Write one consolidated report for this run ---------------------
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(directory, f"Ingest_merge_report_{timestamp}.txt")

    lines = [
        "Ingest merge report",
        f"Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Directory: {directory}",
        source_line,
        "",
        f"New issues added: {added_total} ({new_pages_total} pages)",
        f"Already present (unchanged): {matched_count} ({matched_pages_total} pages)",
        f"Duplicates with differing data (needs review): {len(mismatch_lines)} ({mismatch_pages_total} pages)",
        "",
    ]

    if counts_by_source:
        lines.append("Breakdown by source:")
        for source in sorted(counts_by_source):
            c = counts_by_source[source]
            lines.append(
                f"  - {source}: {c['new']} new ({c['new_pages']} pages), "
                f"{c['matched']} already present ({c['matched_pages']} pages), "
                f"{c['mismatch']} differing ({c['mismatch_pages']} pages)"
            )
        lines.append("")

    if written_files:
        lines.append("Files written/updated:")
        lines.extend(f"  - {path}" for path in written_files)
        lines.append("")

    if mismatch_lines:
        lines.append("DUPLICATES WITH DIFFERING DATA:")
        lines.extend(mismatch_lines)
        lines.append("")

    if volume_problems:
        lines.append("OTHER PROBLEMS (from reading the source files):")
        lines.extend(f"  - {p}" for p in volume_problems)
        lines.append("")

    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print()
    print("Done!")
    print(f"  New issues added: {added_total} ({new_pages_total} pages)")
    print(f"  Already present (unchanged): {matched_count} ({matched_pages_total} pages)")
    print(f"  Duplicates with differing data: {len(mismatch_lines)} ({mismatch_pages_total} pages)")
    if counts_by_source:
        print("  Breakdown by source:")
        for source in sorted(counts_by_source):
            c = counts_by_source[source]
            print(f"    - {source}: {c['new']} new ({c['new_pages']} pages), "
                  f"{c['matched']} already present ({c['matched_pages']} pages), "
                  f"{c['mismatch']} differing ({c['mismatch_pages']} pages)")
    print()
    if written_files:
        print("  Files written/updated:")
        for path in written_files:
            print(f"    - {path}")
    print(f"  Process report: {report_path}")


def run_by_year(directory):
    """
    The default directory-scanning mode: read every tifflist and Town
    List file found anywhere in `directory` and merge them all in - no
    filename pairing required. See the module comment above for the
    merge/dedup rules and file-naming convention.
    """
    print(f"Scanning {directory} for tifflist and Town List files...")
    all_records, problems, tiff_filenames, townlist_filenames = gather_all_records(directory)

    if not tiff_filenames and not townlist_filenames:
        print("No tifflist or Town List files found.")
        return

    print(f"Found {len(tiff_filenames)} tifflist file(s) and "
          f"{len(townlist_filenames)} Town List file(s).")
    print()

    source_line = (
        f"Tifflist files read: {', '.join(tiff_filenames) or '(none)'}\n"
        f"Town List files read: {', '.join(townlist_filenames) or '(none)'}"
    )
    merge_and_write_records(directory, all_records, problems, source_line)


# =============================================================================
# MAIN - this is what runs when you type "python build_issue_csv.py ..."
# =============================================================================

def main():
    args = sys.argv[1:]

    try:
        load_notes()
    except (FileNotFoundError, ValueError) as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    # ---- A single directory argument (with or without the explicit
    # --by-year flag) - the DEFAULT mode: read every Vol_xx pair found in
    # that directory and merge them into start-end-ingest.csv files.
    directory = None
    if len(args) == 1 and os.path.isdir(args[0]):
        directory = args[0]
    elif len(args) == 2 and args[0] == "--by-year":
        directory = args[1]
        if not os.path.isdir(directory):
            print(f"ERROR: '{directory}' is not a directory.")
            sys.exit(1)

    if directory is not None:
        run_by_year(directory)
        return

    # ---- Manual, single-pair mode - for one-off / oddly-named files that
    # don't follow the Vol_xx_tifflist.txt naming convention. Still
    # merges into whatever "start-end-ingest.csv" files already exist in
    # the same folder, exactly like the default mode does.
    if len(args) != 2:
        print("Usage:")
        print("  python build_issue_csv.py <directory>")
        print("      (default: read every Vol_xx pair in <directory> and merge them)")
        print("  python build_issue_csv.py <tiff_directory_or_listing_file> <townlist_file_or_google_sheet_url>")
        print("      (manual mode, for one-off files that don't follow the Vol_xx naming)")
        sys.exit(1)

    tiff_source = args[0]
    townlist_file = args[1]

    # Only check "does this file exist on disk" for LOCAL Town List
    # files - a Google Sheets URL obviously isn't a path on this
    # computer, so that check would always (wrongly) fail for it.
    if not is_google_sheets_url(townlist_file) and not os.path.isfile(townlist_file):
        print(f"ERROR: Town List file not found: {townlist_file}")
        sys.exit(1)

    # ---- Figure out what kind of thing tiff_source is, and read it ----
    if os.path.isdir(tiff_source):
        # A real folder - scan it for .tif files the old way.
        print(f"Scanning TIFF directory: {tiff_source}")
        tiff_issues, unrecognized_items = scan_tiff_directory(tiff_source)
        output_dir = tiff_source
        source_label = f"the '{tiff_source}' directory"

    elif os.path.isfile(tiff_source):
        # A listing file (.csv or .txt) - read filenames out of it
        # instead of browsing an actual folder.
        print(f"Reading TIFF filenames from listing file: {tiff_source}")
        tiff_issues, unrecognized_items = read_tiff_list_file(tiff_source)
        # Output files go into the SAME FOLDER AS THE LISTING FILE, since
        # that's the only local location we actually have.
        output_dir = os.path.dirname(os.path.abspath(tiff_source)) or "."
        source_label = f"the '{tiff_source}' listing file"

    else:
        print(f"ERROR: '{tiff_source}' is not a directory or a file.")
        sys.exit(1)

    print(f"  Found {len(tiff_issues)} issue(s) worth of TIFF pages.")
    if unrecognized_items:
        noun = "entry" if len(unrecognized_items) == 1 else "entries"
        print(f"  ({len(unrecognized_items)} {noun} didn't match the expected "
              f"naming pattern and will be skipped - see the process report.)")

    print(f"Reading Town List: {townlist_file}")
    grid = read_grid(townlist_file)
    raw_town_lists = parse_town_list(grid)
    town_lists, resolve_problems = resolve_town_list_volumes(
        raw_town_lists, tiff_issues, filename_hint=os.path.basename(townlist_file)
    )
    print(f"  Found {len(town_lists)} issue(s) in the Town List.")

    print("Merging into output files...")
    records, problems = records_from_single_source(
        tiff_issues, town_lists, unrecognized_items, source_label=source_label,
    )
    problems.extend(f"[{townlist_file}] {p}" for p in resolve_problems)
    source_line = f"Source: {tiff_source} + {townlist_file}"
    merge_and_write_records(output_dir, records, problems, source_line)


if __name__ == "__main__":
    main()
